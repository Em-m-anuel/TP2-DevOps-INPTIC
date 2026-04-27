"""
Gestion Etudiants API — INPTIC DevOps TP2
Flask + SQLite + Prometheus + Import CSV/Excel
"""
import os
import io
import csv
import time
import logging
from datetime import datetime
from flask import Flask, jsonify, request, render_template
from flask_sqlalchemy import SQLAlchemy
from prometheus_client import Counter, Histogram, Gauge, generate_latest, CONTENT_TYPE_LATEST

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
START_TIME = time.time()

# ─── Base de donnees ──────────────────────────────────────────────────────────
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////data/students.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'check_same_thread': False},
    'pool_pre_ping': True,
}
db = SQLAlchemy(app)

class Student(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    nom        = db.Column(db.String(100), nullable=False)
    prenom     = db.Column(db.String(100), nullable=False)
    filiere    = db.Column(db.String(50),  nullable=False)
    note       = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id, 'nom': self.nom, 'prenom': self.prenom,
            'filiere': self.filiere, 'note': self.note,
            'created_at': self.created_at.isoformat()
        }

# ─── Metriques Prometheus ─────────────────────────────────────────────────────
REQUEST_COUNT   = Counter('http_requests_total', 'Total HTTP', ['method', 'endpoint', 'status'])
REQUEST_LATENCY = Histogram('http_request_duration_seconds', 'Latence', ['method', 'endpoint'],
                            buckets=[0.01, 0.05, 0.1, 0.25, 0.5, 1.0, 2.5])
IN_PROGRESS     = Gauge('http_requests_inprogress', 'Requetes en cours')

STUDENTS_TOTAL      = Gauge('students_total', 'Total etudiants')
STUDENTS_AVG        = Gauge('students_average_grade', 'Moyenne generale')
STUDENTS_MAX        = Gauge('students_max_grade', 'Note maximale')
STUDENTS_MIN        = Gauge('students_min_grade', 'Note minimale')
STUDENTS_ABOVE_AVG  = Gauge('students_above_average_count', 'Etudiants >= 10')
STUDENTS_BY_FILIERE = Gauge('students_by_filiere', 'Par filiere', ['filiere'])
STUDENT_OPS         = Counter('student_operations_total', 'Operations CRUD', ['operation'])
IMPORT_OPS          = Counter('student_imports_total', 'Imports en lot', ['format'])

def update_metrics():
    students = Student.query.all()
    total = len(students)
    STUDENTS_TOTAL.set(total)
    if total > 0:
        notes = [s.note for s in students]
        STUDENTS_AVG.set(round(sum(notes) / total, 2))
        STUDENTS_MAX.set(max(notes))
        STUDENTS_MIN.set(min(notes))
        STUDENTS_ABOVE_AVG.set(sum(1 for n in notes if n >= 10))
        filieres = {}
        for s in students:
            filieres[s.filiere] = filieres.get(s.filiere, 0) + 1
        for f, c in filieres.items():
            STUDENTS_BY_FILIERE.labels(filiere=f).set(c)
    else:
        for g in [STUDENTS_AVG, STUDENTS_MAX, STUDENTS_MIN, STUDENTS_ABOVE_AVG]:
            g.set(0)

# ─── Middleware ───────────────────────────────────────────────────────────────
@app.before_request
def before():
    request.start_time = time.time()
    IN_PROGRESS.inc()

@app.after_request
def after(response):
    resp_time = time.time() - request.start_time
    REQUEST_LATENCY.labels(method=request.method, endpoint=request.path).observe(resp_time)
    REQUEST_COUNT.labels(method=request.method, endpoint=request.path,
                         status=response.status_code).inc()
    IN_PROGRESS.dec()
    return response

# ─── Pages web ────────────────────────────────────────────────────────────────
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/students')
def students_page():
    return render_template('students.html')

# ─── API ──────────────────────────────────────────────────────────────────────
@app.route('/health')
def health():
    update_metrics()
    return jsonify({
        "status": "healthy",
        "uptime_seconds": round(time.time() - START_TIME, 2),
        "timestamp": datetime.now().isoformat(),
        "students_count": Student.query.count()
    }), 200

@app.route('/metrics')
def metrics():
    update_metrics()
    return generate_latest(), 200, {'Content-Type': CONTENT_TYPE_LATEST}

@app.route('/api/students', methods=['GET'])
def get_students():
    filiere = request.args.get('filiere')
    q = Student.query.filter_by(filiere=filiere) if filiere else Student.query
    rows = q.order_by(Student.id).all()
    return jsonify({"count": len(rows), "students": [s.to_dict() for s in rows]})

@app.route('/api/students/<int:sid>', methods=['GET'])
def get_student(sid):
    return jsonify(Student.query.get_or_404(sid).to_dict())

@app.route('/api/students', methods=['POST'])
def create_student():
    data = request.get_json() or {}
    if not all(k in data for k in ['nom', 'prenom', 'filiere']):
        return jsonify({"error": "Champs requis : nom, prenom, filiere"}), 400
    s = Student(nom=data['nom'], prenom=data['prenom'],
                filiere=data['filiere'], note=float(data.get('note', 0.0)))
    db.session.add(s)
    db.session.commit()
    STUDENT_OPS.labels(operation='create').inc()
    update_metrics()
    return jsonify(s.to_dict()), 201

@app.route('/api/students/<int:sid>', methods=['PUT'])
def update_student(sid):
    s = Student.query.get_or_404(sid)
    data = request.get_json() or {}
    for field in ['nom', 'prenom', 'filiere']:
        if field in data: setattr(s, field, data[field])
    if 'note' in data:
        s.note = float(data['note'])
    db.session.commit()
    STUDENT_OPS.labels(operation='update').inc()
    update_metrics()
    return jsonify(s.to_dict())

@app.route('/api/students/<int:sid>', methods=['DELETE'])
def delete_student(sid):
    s = Student.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    STUDENT_OPS.labels(operation='delete').inc()
    update_metrics()
    return jsonify({"message": "Supprime", "id": sid})

@app.route('/api/stats')
def get_stats():
    students = Student.query.all()
    total = len(students)
    if total == 0:
        return jsonify({"total": 0, "moyenne": 0, "max": 0, "min": 0,
                        "above_average": 0, "by_filiere": {}})
    notes = [s.note for s in students]
    filieres = {}
    for s in students:
        filieres[s.filiere] = filieres.get(s.filiere, 0) + 1
    return jsonify({
        "total": total, "moyenne": round(sum(notes)/total, 2),
        "max": max(notes), "min": min(notes),
        "above_average": sum(1 for n in notes if n >= 10),
        "by_filiere": filieres
    })

# ─── Import CSV ───────────────────────────────────────────────────────────────
@app.route('/api/students/import/csv', methods=['POST'])
def import_csv():
    try:
        if 'file' in request.files:
            content = request.files['file'].read().decode('utf-8')
        elif request.content_type and 'text/csv' in request.content_type:
            content = request.data.decode('utf-8')
        else:
            return jsonify({"error": "Envoie un fichier CSV"}), 400

        reader = csv.DictReader(io.StringIO(content))
        imported, errors = [], []

        for i, row in enumerate(reader, 1):
            try:
                nom    = row.get('nom', '').strip()
                prenom = row.get('prenom', '').strip()
                filiere = row.get('filiere', '').strip()
                note   = float(row.get('note', 0) or 0)

                if not nom or not prenom or not filiere:
                    errors.append(f"Ligne {i+1} : donnees manquantes")
                    continue

                s = Student(nom=nom, prenom=prenom, filiere=filiere, note=note)
                db.session.add(s)
                imported.append({'nom': nom, 'prenom': prenom, 'filiere': filiere, 'note': note})
                STUDENT_OPS.labels(operation='create').inc()
            except Exception as e:
                errors.append(f"Ligne {i+1} : {str(e)}")

        db.session.commit()
        update_metrics()
        IMPORT_OPS.labels(format='csv').inc()
        return jsonify({"imported": len(imported), "errors": len(errors)}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# ─── Import Excel ─────────────────────────────────────────────────────────────
@app.route('/api/students/import/excel', methods=['POST'])
def import_excel():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl non installe"}), 500

    if 'file' not in request.files:
        return jsonify({"error": "Fichier manquant"}), 400

    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active
        headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
        idx = {h: headers.index(h) for h in headers if h}
        imported = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            s = Student(nom=str(row[idx['nom']]), prenom=str(row[idx['prenom']]), 
                        filiere=str(row[idx['filiere']]), note=float(row[idx.get('note', 0)] or 0))
            db.session.add(s)
            imported += 1
            STUDENT_OPS.labels(operation='create').inc()

        db.session.commit()
        update_metrics()
        IMPORT_OPS.labels(format='excel').inc()
        return jsonify({"imported": imported}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# ─── Export CSV ───────────────────────────────────────────────────────────────
@app.route('/api/students/export/csv', methods=['GET'])
def export_csv():
    students = Student.query.order_by(Student.filiere, Student.nom).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['id', 'nom', 'prenom', 'filiere', 'note', 'created_at'])
    for s in students:
        writer.writerow([s.id, s.nom, s.prenom, s.filiere, s.note, s.created_at])
    return output.getvalue(), 200, {'Content-Type': 'text/csv'}

# ─── Init base ────────────────────────────────────────────────────────────────
def init_db():
    os.makedirs('/data', exist_ok=True)
    with app.app_context():
        # Flask-SQLAlchemy 3.x : utiliser metadata directement avec checkfirst
        db.metadata.create_all(bind=db.engine, checkfirst=True)
        if Student.query.count() == 0:
            seeds = [
                Student(nom="G",      prenom="DesmonD", filiere="LP-DAR",  note=15.5),
                Student(nom="Camara", prenom="Ibrahim", filiere="LP-INFO", note=13.5),
                Student(nom="Bah",    prenom="Mariama", filiere="LP-DAR",  note=17.0),
                Student(nom="Diallo", prenom="Aissatou",filiere="LP-INFO", note=11.5),
            ]
            for s in seeds:
                db.session.add(s)
            db.session.commit()
            logger.info("Base initialisee avec %d etudiants", len(seeds))

init_db()
